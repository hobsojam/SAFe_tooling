from pathlib import Path

import openpyxl
import openpyxl.utils
import typer
from openpyxl.styles import Alignment, Font, PatternFill
from rich.console import Console
from rich.table import Table

import safe.cli.state as state
from safe.store.db import get_db
from safe.store.repos import get_repos

backlog_app = typer.Typer(help="View and manage the program backlog")
console = Console()

_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_ALIGN = Alignment(horizontal="center")
_COL_WIDTHS = [6, 35, 15, 15, 15, 6, 6, 8, 8]


def _repos():
    return get_repos(get_db(state.db_path) if state.db_path else None)


@backlog_app.command("show")
def backlog_show(
    pi_id: str | None = typer.Option(None, "--pi-id", help="Filter by PI"),
):
    """Show program backlog ranked by WSJF score."""
    repos = _repos()
    features = repos.features.find(pi_id=pi_id) if pi_id else repos.features.get_all()
    if not features:
        console.print("Backlog is empty.")
        return
    features.sort(key=lambda f: f.wsjf_score, reverse=True)

    team_map = {t.id: t.name for t in repos.teams.get_all()}
    pi_map = {p.id: p.name for p in repos.pis.get_all()}

    table = Table("#", "Name", "Status", "CoD", "Size", "WSJF", "Stories", "Team", "PI")
    for rank, f in enumerate(features, 1):
        story_count = str(len(repos.stories.find(feature_id=f.id)))
        table.add_row(
            str(rank),
            f.name,
            f.status.value,
            str(f.cost_of_delay),
            str(f.job_size),
            str(f.wsjf_score),
            story_count,
            team_map.get(f.team_id, f.team_id) if f.team_id else "-",
            pi_map.get(f.pi_id, f.pi_id) if f.pi_id else "-",
        )
    console.print(table)


@backlog_app.command("export")
def backlog_export(
    pi_id: str | None = typer.Option(None, "--pi-id", help="Filter by PI"),
    output: Path = typer.Option(Path("backlog.xlsx"), "--output", "-o", help="Output .xlsx file"),
):
    """Export WSJF-ranked backlog to an Excel workbook."""
    repos = _repos()
    features = repos.features.find(pi_id=pi_id) if pi_id else repos.features.get_all()
    if not features:
        console.print("Backlog is empty.")
        return
    features.sort(key=lambda f: f.wsjf_score, reverse=True)

    team_map = {t.id: t.name for t in repos.teams.get_all()}
    pi_map = {p.id: p.name for p in repos.pis.get_all()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Backlog"

    headers = ["Rank", "Feature", "PI", "Team", "Status", "CoD", "Size", "WSJF", "Stories"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN

    ws.freeze_panes = "A2"

    for rank, f in enumerate(features, 1):
        story_count = len(repos.stories.find(feature_id=f.id))
        ws.append(
            [
                rank,
                f.name,
                pi_map.get(f.pi_id, f.pi_id) if f.pi_id else "-",
                team_map.get(f.team_id, f.team_id) if f.team_id else "-",
                f.status.value,
                f.cost_of_delay,
                f.job_size,
                f.wsjf_score,
                story_count,
            ]
        )

    for i, width in enumerate(_COL_WIDTHS, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    wb.save(output)
    console.print(f"Exported {len(features)} feature(s) to [bold]{output}[/bold]")
