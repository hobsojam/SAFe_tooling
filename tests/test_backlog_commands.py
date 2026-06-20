from io import StringIO
from pathlib import Path

import openpyxl
import pytest
from rich.console import Console
from typer.testing import CliRunner

import safe.cli.backlog as backlog_module
import safe.cli.feature as feature_module
import safe.cli.state as state
from safe.cli.main import app
from safe.store.db import get_db
from safe.store.repos import get_repos

runner = CliRunner()

FEATURE_ARGS = [
    "--name",
    "Auth Service",
    "--user-value",
    "8",
    "--time-crit",
    "5",
    "--risk-reduction",
    "3",
    "--job-size",
    "4",
]


@pytest.fixture(autouse=True)
def reset_state():
    state.db_path = None
    yield
    state.db_path = None


@pytest.fixture
def db_path(tmp_path: Path) -> Path:
    return tmp_path / "test.json"


@pytest.fixture(autouse=True)
def patch_console(monkeypatch):
    buf = StringIO()
    test_console = Console(file=buf, highlight=False, markup=False, width=200)
    monkeypatch.setattr(backlog_module, "console", test_console)
    yield buf


def invoke(db_path: Path, *args):
    return runner.invoke(app, ["--db-path", str(db_path)] + list(args))


def repos_for(db_path: Path):
    return get_repos(get_db(db_path))


class TestBacklogShow:
    def test_empty(self, db_path, patch_console):
        invoke(db_path, "backlog", "show")
        assert "empty" in patch_console.getvalue().lower()

    def test_shows_feature(self, db_path, patch_console):
        invoke(db_path, "feature", "add", *FEATURE_ARGS)
        patch_console.truncate(0)
        patch_console.seek(0)
        invoke(db_path, "backlog", "show")
        assert "Auth Service" in patch_console.getvalue()

    def test_ranked_by_wsjf(self, db_path, patch_console):
        invoke(
            db_path,
            "feature",
            "add",
            "--name",
            "Low",
            "--user-value",
            "1",
            "--time-crit",
            "1",
            "--risk-reduction",
            "1",
            "--job-size",
            "13",
        )
        invoke(db_path, "feature", "add", *FEATURE_ARGS)
        patch_console.truncate(0)
        patch_console.seek(0)
        invoke(db_path, "backlog", "show")
        output = patch_console.getvalue()
        assert output.index("Auth Service") < output.index("Low")

    def test_exit_code_success(self, db_path, patch_console):
        result = invoke(db_path, "backlog", "show")
        assert result.exit_code == 0


class TestBacklogShowTeamResolution:
    def test_shows_team_name_not_uuid(self, db_path, patch_console):
        invoke(db_path, "team", "create", "--name", "Alpha", "--members", "6")
        team_id = repos_for(db_path).teams.get_all()[0].id
        invoke(db_path, "feature", "add", *FEATURE_ARGS)
        feature_id = repos_for(db_path).features.get_all()[0].id
        invoke(db_path, "feature", "assign", feature_id, "--team-id", team_id)
        patch_console.truncate(0)
        patch_console.seek(0)
        invoke(db_path, "backlog", "show")
        output = patch_console.getvalue()
        assert "Alpha" in output
        assert team_id not in output


class TestBacklogExport:
    def test_empty_backlog_prints_message(self, db_path, patch_console, tmp_path):
        out = tmp_path / "out.xlsx"
        result = invoke(db_path, "backlog", "export", "--output", str(out))
        assert result.exit_code == 0
        assert "empty" in patch_console.getvalue().lower()
        assert not out.exists()

    def test_creates_xlsx_file(self, db_path, patch_console, tmp_path):
        invoke(db_path, "feature", "add", *FEATURE_ARGS)
        out = tmp_path / "out.xlsx"
        result = invoke(db_path, "backlog", "export", "--output", str(out))
        assert result.exit_code == 0
        assert out.exists()

    def test_xlsx_has_header_row(self, db_path, patch_console, tmp_path):
        invoke(db_path, "feature", "add", *FEATURE_ARGS)
        out = tmp_path / "out.xlsx"
        invoke(db_path, "backlog", "export", "--output", str(out))
        ws = openpyxl.load_workbook(out).active
        headers = [cell.value for cell in ws[1]]
        assert "Feature" in headers
        assert "WSJF" in headers
        assert "CoD" in headers

    def test_feature_appears_in_rows(self, db_path, patch_console, tmp_path):
        invoke(db_path, "feature", "add", *FEATURE_ARGS)
        out = tmp_path / "out.xlsx"
        invoke(db_path, "backlog", "export", "--output", str(out))
        ws = openpyxl.load_workbook(out).active
        all_values = [cell.value for row in ws.iter_rows(min_row=2) for cell in row]
        assert "Auth Service" in all_values

    def test_wsjf_column_is_numeric(self, db_path, patch_console, tmp_path):
        invoke(db_path, "feature", "add", *FEATURE_ARGS)
        out = tmp_path / "out.xlsx"
        invoke(db_path, "backlog", "export", "--output", str(out))
        ws = openpyxl.load_workbook(out).active
        headers = [cell.value for cell in ws[1]]
        wsjf_col = headers.index("WSJF") + 1
        wsjf_value = ws.cell(row=2, column=wsjf_col).value
        assert isinstance(wsjf_value, (int, float))

    def test_team_name_resolved_in_export(self, db_path, patch_console, tmp_path):
        invoke(db_path, "team", "create", "--name", "Alpha", "--members", "6")
        team_id = repos_for(db_path).teams.get_all()[0].id
        invoke(db_path, "feature", "add", *FEATURE_ARGS)
        feature_id = repos_for(db_path).features.get_all()[0].id
        invoke(db_path, "feature", "assign", feature_id, "--team-id", team_id)
        out = tmp_path / "out.xlsx"
        invoke(db_path, "backlog", "export", "--output", str(out))
        ws = openpyxl.load_workbook(out).active
        all_values = [cell.value for row in ws.iter_rows(min_row=2) for cell in row]
        assert "Alpha" in all_values
        assert team_id not in all_values

    def test_pi_filter(self, db_path, patch_console, tmp_path):
        invoke(db_path, "art", "create", "--name", "ART1")
        art_id = repos_for(db_path).arts.get_all()[0].id
        invoke(
            db_path,
            "pi",
            "create",
            "--name",
            "PI1",
            "--art-id",
            art_id,
            "--start",
            "2026-01-05",
            "--end",
            "2026-03-27",
        )
        pi_id = repos_for(db_path).pis.get_all()[0].id
        invoke(db_path, "feature", "add", *FEATURE_ARGS, "--pi-id", pi_id)
        invoke(
            db_path,
            "feature",
            "add",
            "--name",
            "Other",
            "--user-value",
            "1",
            "--time-crit",
            "1",
            "--risk-reduction",
            "1",
            "--job-size",
            "1",
        )
        out = tmp_path / "out.xlsx"
        invoke(db_path, "backlog", "export", "--pi-id", pi_id, "--output", str(out))
        ws = openpyxl.load_workbook(out).active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) == 1
        assert rows[0][1] == "Auth Service"

    def test_ranked_by_wsjf_descending(self, db_path, patch_console, tmp_path):
        invoke(
            db_path,
            "feature",
            "add",
            "--name",
            "Low",
            "--user-value",
            "1",
            "--time-crit",
            "1",
            "--risk-reduction",
            "1",
            "--job-size",
            "13",
        )
        invoke(db_path, "feature", "add", *FEATURE_ARGS)
        out = tmp_path / "out.xlsx"
        invoke(db_path, "backlog", "export", "--output", str(out))
        ws = openpyxl.load_workbook(out).active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert rows[0][1] == "Auth Service"
        assert rows[1][1] == "Low"

    def test_prints_confirmation(self, db_path, patch_console, tmp_path):
        invoke(db_path, "feature", "add", *FEATURE_ARGS)
        out = tmp_path / "out.xlsx"
        invoke(db_path, "backlog", "export", "--output", str(out))
        assert str(out) in patch_console.getvalue()


class TestWsjfRankAlias:
    @pytest.fixture
    def feature_buf(self, monkeypatch):
        buf = StringIO()
        monkeypatch.setattr(
            feature_module, "console", Console(file=buf, highlight=False, markup=False, width=200)
        )
        return buf

    def test_rank_alias_exit_code(self, db_path, patch_console, feature_buf):
        result = invoke(db_path, "wsjf", "rank")
        assert result.exit_code == 0

    def test_rank_alias_empty(self, db_path, patch_console, feature_buf):
        invoke(db_path, "wsjf", "rank")
        assert "No features found" in feature_buf.getvalue()

    def test_rank_alias_shows_feature(self, db_path, patch_console, feature_buf):
        invoke(db_path, "feature", "add", *FEATURE_ARGS)
        feature_buf.truncate(0)
        feature_buf.seek(0)
        invoke(db_path, "wsjf", "rank")
        assert "Auth Service" in feature_buf.getvalue()


class TestStoryClearIteration:
    def test_clear_iteration_with_empty_string(self, db_path, patch_console):
        invoke(db_path, "feature", "add", *FEATURE_ARGS)
        invoke(db_path, "team", "create", "--name", "Alpha", "--members", "6")
        repos = repos_for(db_path)
        fid = repos.features.get_all()[0].id
        tid = repos.teams.get_all()[0].id
        invoke(db_path, "art", "create", "--name", "ART")
        art_id = repos_for(db_path).arts.get_all()[0].id
        invoke(
            db_path,
            "pi",
            "create",
            "--name",
            "PI 1",
            "--art-id",
            art_id,
            "--start",
            "2026-01-05",
            "--end",
            "2026-03-27",
        )
        pi_id = repos_for(db_path).pis.get_all()[0].id
        invoke(
            db_path,
            "pi",
            "iteration",
            "add",
            "--pi-id",
            pi_id,
            "--number",
            "1",
            "--start",
            "2026-01-05",
            "--end",
            "2026-01-16",
        )
        iter_id = repos_for(db_path).iterations.get_all()[0].id
        invoke(
            db_path,
            "story",
            "add",
            "--name",
            "S1",
            "--feature-id",
            fid,
            "--team-id",
            tid,
            "--points",
            "3",
            "--iteration-id",
            iter_id,
        )
        story = repos_for(db_path).stories.get_all()[0]
        assert story.iteration_id == iter_id
        invoke(db_path, "story", "update", story.id, "--iteration-id", "")
        updated = repos_for(db_path).stories.get(story.id)
        assert updated.iteration_id is None
